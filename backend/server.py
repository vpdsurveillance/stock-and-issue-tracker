from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import logging
import uuid
import bcrypt
import jwt as pyjwt
from datetime import datetime, timezone, timedelta, date
from typing import Optional, List, Literal
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from motor.motor_asyncio import AsyncIOMotorClient
from starlette.middleware.cors import CORSMiddleware
from openpyxl import Workbook, load_workbook

# --- Config ---
JWT_SECRET = os.environ.get("JWT_SECRET", "changeme")
JWT_ALG = "HS256"
DEPARTMENTS = ("MDS", "VPD", "Media")
Department = Literal["MDS", "VPD", "Media"]

# --- DB ---
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("stockapp")


# --- Helpers ---
def hash_password(pwd: str) -> str:
    return bcrypt.hashpw(pwd.encode(), bcrypt.gensalt()).decode()


def verify_password(pwd: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pwd.encode(), hashed.encode())
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=12),
        "type": "access",
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        h = request.headers.get("Authorization", "")
        if h.startswith("Bearer "):
            token = h[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"id": payload["sub"]})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    user.pop("_id", None)
    user.pop("password_hash", None)
    return user


async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return user


def _iso(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).isoformat()


def _parse_date(s: str) -> datetime:
    """Parse an ISO date/datetime string into aware datetime (UTC)."""
    if isinstance(s, datetime):
        return s if s.tzinfo else s.replace(tzinfo=timezone.utc)
    if isinstance(s, date):
        return datetime(s.year, s.month, s.day, tzinfo=timezone.utc)
    dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


# --- Models ---
class LoginBody(BaseModel):
    email: EmailStr
    password: str


class UserOut(BaseModel):
    id: str
    email: EmailStr
    name: str
    role: str


class ItemIn(BaseModel):
    name: str
    pack_size: str
    department: Department


class ItemOut(ItemIn):
    id: str
    created_at: str


class StockEntryIn(BaseModel):
    item_id: str
    department: Department
    item_name: str
    pack_size: str
    quantity: int = Field(gt=0)
    receipt_date: str  # ISO date
    lot_number: str
    expiry_date: str  # ISO date
    manufacturer: str
    supplier: str
    program: str


class IssueIn(BaseModel):
    item_id: str
    department: Department
    item_name: str
    pack_size: str
    expiry_date: str
    quantity: int = Field(gt=0)
    issued_section: str
    issue_date: Optional[str] = None
    program: Optional[str] = ""
    lot_number: Optional[str] = ""


class IssueBatchIn(BaseModel):
    items: List[IssueIn]


# --- Auth endpoints ---
@api.post("/auth/login")
async def login(body: LoginBody, response: Response):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(user["id"], user["email"], user["role"])
    response.set_cookie("access_token", token, httponly=True, secure=False,
                        samesite="lax", max_age=43200, path="/")
    return {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"], "token": token}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


# --- Items ---
@api.get("/items")
async def list_items(department: Optional[str] = None, search: Optional[str] = None,
                     user: dict = Depends(get_current_user)):
    q = {}
    if department:
        q["department"] = department
    if search:
        q["name"] = {"$regex": search, "$options": "i"}
    docs = await db.items.find(q, {"_id": 0}).sort("name", 1).to_list(2000)
    return docs


@api.post("/items")
async def create_item(body: ItemIn, user: dict = Depends(require_admin)):
    if body.department not in DEPARTMENTS:
        raise HTTPException(400, "Invalid department")
    exists = await db.items.find_one({"department": body.department,
                                      "name": {"$regex": f"^{body.name}$", "$options": "i"},
                                      "pack_size": body.pack_size})
    if exists:
        raise HTTPException(400, "Item with same name and pack size already exists in this department")
    doc = {
        "id": str(uuid.uuid4()),
        "name": body.name.strip(),
        "pack_size": body.pack_size.strip(),
        "department": body.department,
        "created_at": _iso(datetime.now(timezone.utc)),
    }
    await db.items.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.delete("/items/{item_id}")
async def delete_item(item_id: str, user: dict = Depends(require_admin)):
    r = await db.items.delete_one({"id": item_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"ok": True}


# --- Meta autocomplete ---
@api.get("/meta/{field}")
async def meta_field(field: str, user: dict = Depends(get_current_user)):
    if field not in ("manufacturers", "suppliers", "programs", "sections"):
        raise HTTPException(400, "Invalid field")
    src_field = {"manufacturers": "manufacturer", "suppliers": "supplier",
                 "programs": "program", "sections": "issued_section"}[field]
    coll = db.issues if field == "sections" else db.stock_entries
    vals = await coll.distinct(src_field)
    return sorted([v for v in vals if v])


# --- Stock entries ---
@api.post("/stock")
async def create_stock(body: StockEntryIn, user: dict = Depends(get_current_user)):
    item = await db.items.find_one({"id": body.item_id})
    if not item:
        raise HTTPException(404, "Item not found")
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["receipt_date"] = _iso(_parse_date(doc["receipt_date"]))
    doc["expiry_date"] = _iso(_parse_date(doc["expiry_date"]))
    doc["created_at"] = _iso(datetime.now(timezone.utc))
    doc["created_by"] = user["email"]
    await db.stock_entries.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/stock")
async def list_stock(department: Optional[str] = None, search: Optional[str] = None,
                     from_date: Optional[str] = None, to_date: Optional[str] = None,
                     program: Optional[str] = None,
                     user: dict = Depends(get_current_user)):
    q = {}
    if department:
        q["department"] = department
    if search:
        q["item_name"] = {"$regex": search, "$options": "i"}
    if program:
        q["program"] = program
    if from_date or to_date:
        r = {}
        if from_date:
            r["$gte"] = _iso(_parse_date(from_date))
        if to_date:
            r["$lte"] = _iso(_parse_date(to_date) + timedelta(days=1))
        q["receipt_date"] = r
    docs = await db.stock_entries.find(q, {"_id": 0}).sort("receipt_date", -1).to_list(5000)
    return docs


@api.delete("/stock/{sid}")
async def del_stock(sid: str, user: dict = Depends(require_admin)):
    r = await db.stock_entries.delete_one({"id": sid})
    if r.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"ok": True}


# --- Issues ---
@api.post("/issues")
async def create_issue(body: IssueIn, user: dict = Depends(get_current_user)):
    item = await db.items.find_one({"id": body.item_id})
    if not item:
        raise HTTPException(404, "Item not found")
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["expiry_date"] = _iso(_parse_date(doc["expiry_date"]))
    doc["issue_date"] = _iso(_parse_date(doc["issue_date"])) if doc.get("issue_date") else _iso(datetime.now(timezone.utc))
    doc["created_at"] = _iso(datetime.now(timezone.utc))
    doc["created_by"] = user["email"]
    await db.issues.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/issues")
async def list_issues(department: Optional[str] = None, search: Optional[str] = None,
                      from_date: Optional[str] = None, to_date: Optional[str] = None,
                      program: Optional[str] = None, section: Optional[str] = None,
                      user: dict = Depends(get_current_user)):
    q = {}
    if department:
        q["department"] = department
    if search:
        q["item_name"] = {"$regex": search, "$options": "i"}
    if program:
        q["program"] = program
    if section:
        q["issued_section"] = section
    if from_date or to_date:
        r = {}
        if from_date:
            r["$gte"] = _iso(_parse_date(from_date))
        if to_date:
            r["$lte"] = _iso(_parse_date(to_date) + timedelta(days=1))
        q["issue_date"] = r
    docs = await db.issues.find(q, {"_id": 0}).sort("issue_date", -1).to_list(5000)
    return docs


@api.post("/issues/batch")
async def create_issue_batch(body: IssueBatchIn, user: dict = Depends(get_current_user)):
    created = []
    for it in body.items:
        item = await db.items.find_one({"id": it.item_id})
        if not item:
            raise HTTPException(404, f"Item not found: {it.item_name}")
        doc = it.model_dump()
        doc["id"] = str(uuid.uuid4())
        doc["expiry_date"] = _iso(_parse_date(doc["expiry_date"]))
        doc["issue_date"] = _iso(_parse_date(doc["issue_date"])) if doc.get("issue_date") else _iso(datetime.now(timezone.utc))
        doc["created_at"] = _iso(datetime.now(timezone.utc))
        doc["created_by"] = user["email"]
        await db.issues.insert_one(doc)
        doc.pop("_id", None)
        created.append(doc)
    return {"created": len(created), "items": created}


@api.delete("/issues/{iid}")
async def del_issue(iid: str, user: dict = Depends(require_admin)):
    r = await db.issues.delete_one({"id": iid})
    if r.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"ok": True}


# --- Aggregations / Reports ---
async def _current_stock_rows(department: Optional[str] = None, program: Optional[str] = None):
    """Returns list of balance rows grouped by
    (department, item_name, pack_size, expiry_date, lot_number, manufacturer, supplier, program)."""
    q = {}
    if department: q["department"] = department
    if program: q["program"] = program
    entries = await db.stock_entries.find(q, {"_id": 0}).to_list(20000)
    q2 = {}
    if department: q2["department"] = department
    if program: q2["program"] = program
    issues = await db.issues.find(q2, {"_id": 0}).to_list(20000)

    balances = {}
    for e in entries:
        k = (e["department"], e["item_name"], e["pack_size"], e["expiry_date"],
             e.get("lot_number", ""), e.get("manufacturer", ""),
             e.get("supplier", ""), e.get("program", ""))
        b = balances.setdefault(k, {"received": 0, "issued": 0})
        b["received"] += int(e["quantity"])

    # Issues aggregated at (department, item_name, pack_size, expiry_date) level.
    # Deduct against matching lot rows sorted by lot_number (round-robin fill).
    issue_map = {}
    for i in issues:
        k = (i["department"], i["item_name"], i["pack_size"], i["expiry_date"])
        issue_map[k] = issue_map.get(k, 0) + int(i["quantity"])

    for k, qty in issue_map.items():
        # find matching lot rows
        matches = [rk for rk in balances if rk[:4] == k]
        matches.sort()
        remaining = qty
        for rk in matches:
            avail = balances[rk]["received"] - balances[rk]["issued"]
            take = min(avail, remaining)
            balances[rk]["issued"] += take
            remaining -= take
            if remaining <= 0:
                break
        if remaining > 0 and matches:
            # over-issued, attribute rest to first lot
            balances[matches[0]]["issued"] += remaining

    rows = []
    for k, v in balances.items():
        dep, name, pack, exp, lot, mfr, sup, prog = k
        rows.append({
            "department": dep, "item_name": name, "pack_size": pack,
            "expiry_date": exp, "lot_number": lot, "manufacturer": mfr,
            "supplier": sup, "program": prog,
            "received": v["received"], "issued": v["issued"],
            "balance": v["received"] - v["issued"],
        })
    rows.sort(key=lambda r: (r["department"], r["item_name"], r["expiry_date"]))
    return rows


@api.get("/reports/current-stock")
async def current_stock(department: Optional[str] = None, search: Optional[str] = None,
                        program: Optional[str] = None,
                        user: dict = Depends(get_current_user)):
    rows = await _current_stock_rows(department, program)
    if search:
        s = search.lower()
        rows = [r for r in rows if s in r["item_name"].lower()]
    return rows


@api.get("/reports/short-expiry")
async def short_expiry(days: int = 90, department: Optional[str] = None,
                       program: Optional[str] = None, search: Optional[str] = None,
                       user: dict = Depends(get_current_user)):
    rows = await _current_stock_rows(department, program)
    now = datetime.now(timezone.utc)
    cutoff = now + timedelta(days=days)
    out = []
    for r in rows:
        if r["balance"] <= 0:
            continue
        exp = _parse_date(r["expiry_date"])
        if exp <= cutoff:
            r2 = dict(r)
            r2["days_to_expiry"] = (exp - now).days
            out.append(r2)
    out.sort(key=lambda x: x["days_to_expiry"])
    return out


async def _utilisation_by_month(department: Optional[str] = None):
    """Aggregate monthly issues (quantity) per (department, item_name, pack_size)."""
    q = {"department": department} if department else {}
    issues = await db.issues.find(q, {"_id": 0}).to_list(20000)
    data = {}  # key -> {month_str: qty}
    for i in issues:
        d = _parse_date(i["issue_date"])
        m = d.strftime("%Y-%m")
        k = (i["department"], i["item_name"], i["pack_size"])
        row = data.setdefault(k, {})
        row[m] = row.get(m, 0) + int(i["quantity"])
    return data


@api.get("/reports/monthly-utilisation")
async def monthly_util(year: Optional[int] = None, department: Optional[str] = None,
                       program: Optional[str] = None, search: Optional[str] = None,
                       user: dict = Depends(get_current_user)):
    year = year or datetime.now(timezone.utc).year
    q = {}
    if department: q["department"] = department
    if program: q["program"] = program
    issues = await db.issues.find(q, {"_id": 0}).to_list(20000)
    data = {}
    for i in issues:
        d = _parse_date(i["issue_date"])
        m = d.strftime("%Y-%m")
        k = (i["department"], i["item_name"], i["pack_size"])
        row = data.setdefault(k, {})
        row[m] = row.get(m, 0) + int(i["quantity"])
    result = []
    for (dep, name, pack), months in data.items():
        if search and search.lower() not in name.lower():
            continue
        row = {"department": dep, "item_name": name, "pack_size": pack, "total": 0}
        for m in range(1, 13):
            key = f"{year}-{m:02d}"
            v = months.get(key, 0)
            row[f"m{m}"] = v
            row["total"] += v
        if row["total"] > 0:
            result.append(row)
    result.sort(key=lambda r: (r["department"], r["item_name"]))
    return result


@api.get("/reports/indent-next-year")
async def indent_next_year(department: Optional[str] = None, program: Optional[str] = None,
                           search: Optional[str] = None,
                           user: dict = Depends(get_current_user)):
    """Indent = avg monthly utilisation over last 12 months * 12."""
    now = datetime.now(timezone.utc)
    start = now - timedelta(days=365)
    q = {"issue_date": {"$gte": _iso(start)}}
    if department: q["department"] = department
    if program: q["program"] = program
    issues = await db.issues.find(q, {"_id": 0}).to_list(20000)
    agg = {}
    for i in issues:
        k = (i["department"], i["item_name"], i["pack_size"])
        agg[k] = agg.get(k, 0) + int(i["quantity"])
    rows = []
    for (dep, name, pack), total in agg.items():
        if search and search.lower() not in name.lower():
            continue
        rows.append({
            "department": dep, "item_name": name, "pack_size": pack,
            "yearly_utilisation": total,
            "avg_monthly": round(total / 12, 2),
            "indent_next_year": total,
        })
    rows.sort(key=lambda r: (r["department"], -r["indent_next_year"]))
    return rows


async def _critical_value(department: Optional[str] = None):
    """Sum of last 3 full months utilisation per (dep, item_name, pack_size)."""
    now = datetime.now(timezone.utc)
    start = now - timedelta(days=90)
    q = {"issue_date": {"$gte": _iso(start)}}
    if department:
        q["department"] = department
    issues = await db.issues.find(q, {"_id": 0}).to_list(20000)
    agg = {}
    for i in issues:
        k = (i["department"], i["item_name"], i["pack_size"])
        agg[k] = agg.get(k, 0) + int(i["quantity"])
    return agg


@api.get("/reports/low-stock")
async def low_stock(department: Optional[str] = None, program: Optional[str] = None,
                    search: Optional[str] = None,
                    user: dict = Depends(get_current_user)):
    cv = await _critical_value(department)
    rows = await _current_stock_rows(department, program)
    balances = {}
    for r in rows:
        k = (r["department"], r["item_name"], r["pack_size"])
        balances[k] = balances.get(k, 0) + r["balance"]
    out = []
    for k, bal in balances.items():
        c = cv.get(k, 0)
        if search and search.lower() not in k[1].lower():
            continue
        if c > 0 and bal > 0 and c >= bal:
            out.append({"department": k[0], "item_name": k[1], "pack_size": k[2],
                        "balance": bal, "critical_value": c})
    out.sort(key=lambda r: (r["department"], r["item_name"]))
    return out


@api.get("/reports/nil-stock")
async def nil_stock(department: Optional[str] = None, program: Optional[str] = None,
                    search: Optional[str] = None,
                    user: dict = Depends(get_current_user)):
    rows = await _current_stock_rows(department, program)
    balances = {}
    for r in rows:
        k = (r["department"], r["item_name"], r["pack_size"])
        balances[k] = balances.get(k, 0) + r["balance"]
    # Include master items with zero receipts (only if no program filter)
    if not program:
        item_q = {"department": department} if department else {}
        items = await db.items.find(item_q, {"_id": 0}).to_list(2000)
        for it in items:
            k = (it["department"], it["name"], it["pack_size"])
            balances.setdefault(k, 0)
    out = []
    for k, b in balances.items():
        if b > 0: continue
        if search and search.lower() not in k[1].lower():
            continue
        out.append({"department": k[0], "item_name": k[1], "pack_size": k[2], "balance": b})
    out.sort(key=lambda r: (r["department"], r["item_name"]))
    return out


@api.get("/reports/supply-order")
async def supply_order(department: Optional[str] = None, program: Optional[str] = None,
                       search: Optional[str] = None,
                       user: dict = Depends(get_current_user)):
    low = await low_stock(department, program, search, user)
    nil = await nil_stock(department, program, search, user)
    short = await short_expiry(90, department, program, search, user)
    seen = {}
    def key(x): return (x["department"], x["item_name"], x["pack_size"])
    for r in nil:
        k = key(r); seen[k] = {"department": r["department"], "item_name": r["item_name"],
                                "pack_size": r["pack_size"], "balance": r["balance"],
                                "critical_value": 0, "reasons": ["NIL Stock"]}
    for r in low:
        k = key(r)
        if k in seen:
            if "Low Stock" not in seen[k]["reasons"]: seen[k]["reasons"].append("Low Stock")
            seen[k]["critical_value"] = r["critical_value"]
        else:
            seen[k] = {"department": r["department"], "item_name": r["item_name"],
                       "pack_size": r["pack_size"], "balance": r["balance"],
                       "critical_value": r["critical_value"], "reasons": ["Low Stock"]}
    for r in short:
        k = key(r)
        if k in seen:
            if "Short Expiry" not in seen[k]["reasons"]:
                seen[k]["reasons"].append("Short Expiry")
        else:
            seen[k] = {"department": r["department"], "item_name": r["item_name"],
                       "pack_size": r["pack_size"], "balance": r.get("balance", 0),
                       "critical_value": 0, "reasons": ["Short Expiry"]}
    out = list(seen.values())
    out.sort(key=lambda r: (r["department"], r["item_name"]))
    return out


@api.get("/reports/program-consumption")
async def program_consumption(department: Optional[str] = None,
                              user: dict = Depends(get_current_user)):
    """Consumption grouped by program (issued qty). Program comes from stock_entry via
    (item_name, pack_size, expiry_date). If no match, use issue.program field."""
    q = {"department": department} if department else {}
    issues = await db.issues.find(q, {"_id": 0}).to_list(20000)
    entries = await db.stock_entries.find(q, {"_id": 0}).to_list(20000)
    entry_prog = {(e["department"], e["item_name"], e["pack_size"], e["expiry_date"]): e.get("program", "")
                  for e in entries}
    prog_totals = {}
    for i in issues:
        prog = i.get("program") or entry_prog.get(
            (i["department"], i["item_name"], i["pack_size"], i["expiry_date"]), "Unassigned"
        ) or "Unassigned"
        prog_totals[prog] = prog_totals.get(prog, 0) + int(i["quantity"])
    return [{"program": k, "total_issued": v} for k, v in sorted(prog_totals.items(), key=lambda x: -x[1])]


@api.get("/reports/dashboard")
async def dashboard(user: dict = Depends(get_current_user)):
    rows = await _current_stock_rows()
    total_items = await db.items.count_documents({})
    total_entries = await db.stock_entries.count_documents({})
    total_issues = await db.issues.count_documents({})
    total_balance = sum(r["balance"] for r in rows if r["balance"] > 0)
    low = await low_stock(None, None, None, user)
    nil = await nil_stock(None, None, None, user)
    short = await short_expiry(90, None, None, None, user)
    dept_balance = {"MDS": 0, "VPD": 0, "Media": 0}
    for r in rows:
        if r["balance"] > 0 and r["department"] in dept_balance:
            dept_balance[r["department"]] += r["balance"]
    return {
        "total_items": total_items,
        "total_entries": total_entries,
        "total_issues": total_issues,
        "total_balance": total_balance,
        "low_stock_count": len(low),
        "nil_stock_count": len(nil),
        "short_expiry_count": len(short),
        "dept_balance": dept_balance,
    }


# --- Excel Export ---
def _rows_to_xlsx(rows: List[dict], sheet_name: str = "Sheet1") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
    else:
        ws.append(["(no data)"])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


@api.get("/export/{resource}")
async def export_xlsx(resource: str, department: Optional[str] = None,
                     user: dict = Depends(get_current_user)):
    fetch_map = {
        "items": lambda: db.items.find({"department": department} if department else {}, {"_id": 0}).to_list(5000),
        "stock": lambda: db.stock_entries.find({"department": department} if department else {}, {"_id": 0}).sort("receipt_date", -1).to_list(20000),
        "issues": lambda: db.issues.find({"department": department} if department else {}, {"_id": 0}).sort("issue_date", -1).to_list(20000),
    }
    reports = {
        "current-stock": lambda: current_stock(department, None, None, user),
        "monthly-utilisation": lambda: monthly_util(None, department, None, None, user),
        "indent-next-year": lambda: indent_next_year(department, None, None, user),
        "short-expiry": lambda: short_expiry(90, department, None, None, user),
        "low-stock": lambda: low_stock(department, None, None, user),
        "nil-stock": lambda: nil_stock(department, None, None, user),
        "supply-order": lambda: supply_order(department, None, None, user),
    }
    if resource in fetch_map:
        rows = await fetch_map[resource]()
    elif resource in reports:
        rows = await reports[resource]()
    else:
        raise HTTPException(404, "Unknown resource")
    # sanitize
    clean = []
    for r in rows:
        cr = {}
        for k, v in r.items():
            if isinstance(v, list): v = ", ".join(map(str, v))
            cr[k] = v
        clean.append(cr)
    data = _rows_to_xlsx(clean, resource)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{resource}.xlsx"'}
    )


@api.post("/import/items")
async def import_items(department: Department, file: UploadFile = File(...),
                       user: dict = Depends(require_admin)):
    if department not in DEPARTMENTS:
        raise HTTPException(400, "Invalid department")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid xlsx: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"inserted": 0, "skipped": 0}
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    if "name" not in headers or "pack_size" not in headers:
        raise HTTPException(400, "Excel must have 'name' and 'pack_size' columns")
    ni = headers.index("name"); pi = headers.index("pack_size")
    inserted = 0; skipped = 0
    for r in rows[1:]:
        if not r or r[ni] is None: continue
        name = str(r[ni]).strip()
        pack = str(r[pi]).strip() if r[pi] is not None else ""
        if not name: continue
        exists = await db.items.find_one({"department": department,
                                          "name": {"$regex": f"^{name}$", "$options": "i"},
                                          "pack_size": pack})
        if exists: skipped += 1; continue
        await db.items.insert_one({"id": str(uuid.uuid4()), "name": name, "pack_size": pack,
                                   "department": department,
                                   "created_at": _iso(datetime.now(timezone.utc))})
        inserted += 1
    return {"inserted": inserted, "skipped": skipped}


# --- Startup: seed & indexes ---
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.items.create_index([("department", 1), ("name", 1)])
    await db.stock_entries.create_index([("department", 1), ("item_name", 1)])
    await db.issues.create_index([("department", 1), ("item_name", 1)])

    async def _seed(email_env, pw_env, name, role):
        email = os.environ.get(email_env, "").lower()
        pwd = os.environ.get(pw_env, "")
        if not email or not pwd: return
        existing = await db.users.find_one({"email": email})
        if existing is None:
            await db.users.insert_one({
                "id": str(uuid.uuid4()), "email": email, "name": name, "role": role,
                "password_hash": hash_password(pwd),
                "created_at": _iso(datetime.now(timezone.utc)),
            })
            logger.info("Seeded %s user %s", role, email)
        else:
            if not verify_password(pwd, existing["password_hash"]):
                await db.users.update_one({"email": email},
                                          {"$set": {"password_hash": hash_password(pwd)}})
                logger.info("Updated password for %s", email)

    await _seed("ADMIN_EMAIL", "ADMIN_PASSWORD", "Administrator", "admin")
    await _seed("STAFF_EMAIL", "STAFF_PASSWORD", "Staff User", "staff")


@app.on_event("shutdown")
async def shutdown():
    client.close()


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["https://stock-and-issue-tracker-frontend.onrender.com" ],
    allow_methods=["*"],
    allow_headers=["*"],
)
